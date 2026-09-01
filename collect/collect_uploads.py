"""collect 웹 업로드 수집기 — uploads/inbox/의 사용자 업로드 파일을 표준 거래 표로 만든다.

파일 유형별 처리 (단계 문서 collect.md "판단 기준"의 AI 판단/일반 코드 구분 그대로):
- zip: 압축을 풀어 안의 파일을 개별 원천으로 편입 — 일반 코드 (collect.md "하는 단계 0")
- 하나카드 정형 CSV (매핑 규칙 있는 서식): collect.py의 normalize 재사용 — 일반 코드
- 신한법인카드 정형 CSV (매핑 규칙 있는 서식): collect.py의 parse_shinhan 재사용 — 일반 코드.
  이 서식은 항상 법인카드라 결제구분·승인+취소 부호를 매핑 규칙이 확정해서 매긴다 —
  낯선 서식 경로(AI 판단)로 새면 결제구분을 "법인" 문자열 유무로만 추정해 개인결제로 잘못
  기본값이 매겨지고, 승인+취소 쌍도 매핑 규칙 없이 AI가 임의로 처리하게 된다
- 낯선 서식 CSV/TXT/XLSX/PDF(텍스트 레이어 있는 문서만): call-agent.py call_agent_convert_table — AI 판단 (여러 건이면
  한 호출로 묶어 변환 — collect.md "하는 단계 1" 확정, 호출당 고정 비용 절감)
- 영수증·결제 문자 캡처 이미지 (PNG/JPG): call-agent.py call_agent_with_image — AI 판단

계약은 collect.run()과 동일 — collect/result.csv 작성 + {"out_path", "rows", "envelope"} 반환.
지휘(orchestrator/run-pipeline.py)가 웹 실행 시 upload_dir와 함께 부른다.
개별 파일의 AI 처리 실패는 전체를 중단하지 않고 envelope flags에 오류로 남긴다
(collect.md "못 할 때" — 인식 실패 건은 버리지 않는다). zip 안의 개별 항목 실패도 같은 원칙.

사용법(단독 실행): python3 collect/collect_uploads.py 2026-07 [업로드 폴더]
"""

import csv
import importlib.util
import os
import re
import shutil
import sys
import json
import tempfile
import zipfile
from concurrent.futures import ThreadPoolExecutor

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(BASE_DIR)
DEFAULT_UPLOAD_DIR = os.path.join(REPO_ROOT, "uploads", "inbox")
# 현금 등 자동 수집이 어려운 거래의 직접 입력 (collect.md "하는 단계 7") — 웹 서버가
# uploads/inbox/ 밖에 쓴다 (orchestrator/server.py의 POST /uploads/manual)
MANUAL_ENTRIES_PATH = os.path.join(REPO_ROOT, "uploads", "manual_entries.json")

TEXT_EXTS = {".csv", ".txt"}
IMAGE_MEDIA_TYPES = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}

# zip 안에서 허용하는 확장자 — 업로드 화이트리스트(orchestrator/server.py UPLOAD_EXTS)에서 zip 자체만 뺀 것
ZIP_ALLOWED_EXTS = TEXT_EXTS | set(IMAGE_MEDIA_TYPES) | {".xlsx", ".pdf"}
ZIP_MAX_MEMBERS = 50            # 압축 안 파일 개수 상한 (zip bomb 방어)
ZIP_MAX_TOTAL_BYTES = 50 * 1024 * 1024  # 압축 해제 총 용량 상한(50MB, 비압축 기준)

# 하나카드 정형 CSV 판별용 필수 헤더 (collect.py normalize가 읽는 컬럼)
HANA_HEADER = {"이용일자", "이용가맹점", "이용금액", "할부기간", "회차",
               "원금", "수수료", "이용혜택", "혜택금액"}

# 신한법인카드 정형 CSV 판별용 필수 헤더 (collect.py parse_shinhan이 읽는 컬럼) — 이 서식은
# 항상 법인카드이므로 여기서 걸러야 map_table_row의 "법인" 문자열 추정(개인결제 기본값)을
# 타지 않고 parse_shinhan이 결제구분=법인결제·승인+취소 부호 규칙을 그대로 적용한다
SHINHAN_HEADER = {"승인일시", "가맹점명", "승인금액", "통화", "해외이용금액",
                  "승인번호", "사용자", "취소여부"}

AI_WORKERS = 3  # 이미지 병렬 호출 수 (개별 실패는 flags로 흡수 — 낯선 서식은 일괄 1호출)


def _load(name, filename):
    spec = importlib.util.spec_from_file_location(name, os.path.join(BASE_DIR, filename))
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


collect = _load("collect_base", "collect.py")
call_agent = _load("collect_call_agent", "call-agent.py")


def read_text_any(path):
    """업로드 텍스트 파일 디코딩 — 카드사마다 인코딩이 다르다 (KB 명세서는 EUC-KR)."""
    raw = open(path, "rb").read()
    for encoding in ("utf-8-sig", "cp949"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def is_hana_csv(text):
    for line in text.splitlines():
        line = line.strip()
        if line:
            return HANA_HEADER <= {col.strip() for col in line.split(",")}
    return False


def is_shinhan_csv(text):
    for line in text.splitlines():
        line = line.strip()
        if line:
            return SHINHAN_HEADER <= {col.strip() for col in line.split(",")}
    return False


def _is_path_traversal(member_name):
    """경로 탈출(`../`, 절대경로) 항목인지 판별한다."""
    normalized = os.path.normpath(member_name)
    return normalized.startswith("..") or os.path.isabs(normalized)


def extract_zip(zip_path, extract_root):
    """zip 안 항목을 검사해 허용된 확장자만 extract_root 아래에 풀고, 나머지는 오류로 남긴다.

    반환: (extracted_paths, errors). 손상된 zip이거나 상한(ZIP_MAX_MEMBERS·ZIP_MAX_TOTAL_BYTES)을
    넘으면 extracted_paths=[]와 오류 메시지 하나만 돌려주고 zip 전체를 건너뛴다 (collect.md
    "하는 단계 0" 확정 — zip bomb 방어). 개별 항목의 확장자 불허·경로 탈출·zip 안의 zip·
    macOS 부산물(`__MACOSX/`, `.DS_Store`)은 그 항목만 건너뛰고 오류로 남기며 나머지는 처리한다.
    """
    zip_name = os.path.basename(zip_path)
    errors = []
    extracted = []
    try:
        with zipfile.ZipFile(zip_path) as zf:
            infos = [i for i in zf.infolist() if not i.is_dir()]
            if len(infos) > ZIP_MAX_MEMBERS:
                return [], [f"{zip_name}: 압축 안 파일이 {ZIP_MAX_MEMBERS}개를 넘어 처리하지 않음"]
            if sum(i.file_size for i in infos) > ZIP_MAX_TOTAL_BYTES:
                return [], [f"{zip_name}: 압축 해제 용량이 상한"
                            f"({ZIP_MAX_TOTAL_BYTES // (1024 * 1024)}MB)을 넘어 처리하지 않음"]
            dest_dir = os.path.join(extract_root, os.path.splitext(zip_name)[0])
            os.makedirs(dest_dir, exist_ok=True)
            for info in infos:
                name = info.filename
                base = os.path.basename(name.rstrip("/"))
                if name.startswith("__MACOSX/") or base == ".DS_Store":
                    errors.append(f"{zip_name}: {name} — macOS 부산물, 건너뜀")
                    continue
                if _is_path_traversal(name):
                    errors.append(f"{zip_name}: {name} — 경로 탈출 항목, 건너뜀")
                    continue
                ext = os.path.splitext(base)[1].lower()
                if ext == ".zip":
                    errors.append(f"{zip_name}: {name} — zip 안의 zip은 지원하지 않음, 건너뜀")
                    continue
                if ext not in ZIP_ALLOWED_EXTS:
                    errors.append(f"{zip_name}: {name} — 허용되지 않는 확장자({ext or '없음'}), 건너뜀")
                    continue
                dest_path = os.path.join(dest_dir, f"{len(extracted)}_{base}")  # 이름 충돌 방지
                with zf.open(info) as src, open(dest_path, "wb") as out:
                    shutil.copyfileobj(src, out)
                extracted.append(dest_path)
    except zipfile.BadZipFile:
        return [], [f"{zip_name}: 손상된 zip 파일"]
    return extracted, errors


def xlsx_to_text(path):
    """엑셀을 탭 구분 텍스트로 펼친다 — 낯선 서식 변환(AI 판단)의 입력으로 쓴다."""
    from openpyxl import load_workbook  # merge 단계와 같은 기존 의존성
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def pdf_to_text(path):
    """PDF를 텍스트로 뽑는다 — 낯선 서식 변환(AI 판단)의 입력으로 쓴다.

    카드사 명세서 PDF처럼 텍스트 레이어가 있는 문서만 대상 — 스캔 이미지 PDF는
    텍스트가 안 뽑혀 빈 문자열이 되고, process_table_batch가 "내용이 비어 있음" 오류로 남긴다.
    """
    from pypdf import PdfReader
    reader = PdfReader(path)
    return "\n".join(page.extract_text() or "" for page in reader.pages)


# 결제처·비고의 해외결제 표기 승격용 패턴 — "29.99 USD" / "USD 29.99" (단계 문서 "하는 단계 1" 확정).
# 대상 통화는 verify2.md 상식 환율 범위 표와 맞춘다
FOREIGN_CODES = "USD|EUR|JPY|CNY|GBP"
FOREIGN_PATTERNS = [
    re.compile(rf"([0-9]+(?:\.[0-9]+)?)\s*({FOREIGN_CODES})\b"),
    re.compile(rf"\b({FOREIGN_CODES})\s*([0-9]+(?:\.[0-9]+)?)"),
]


def promote_foreign(row):
    """결제처·비고의 통화 표기를 원거래통화·원거래금액 컬럼으로 승격한다 (일반 코드).

    이미 채워진 값은 건드리지 않는다. 통화 컬럼이 없는 원본(가맹점명에 'AMAZON.COM 29.99 USD'처럼
    섞인 경우)을 위한 안전망 — 기간·금액 검증의 환산 검증이 이 컬럼을 근거로 돈다.
    """
    if str(row.get("원거래통화") or "").strip():
        return
    text = f"{row.get('결제처') or ''} {row.get('비고') or ''}"
    for pattern in FOREIGN_PATTERNS:
        m = pattern.search(text)
        if not m:
            continue
        a, b = m.group(1), m.group(2)
        currency, amount = (a, b) if a.isalpha() else (b, a)
        row["원거래통화"] = currency
        row["원거래금액"] = amount
        return


def _norm_for_match(text):
    return re.sub(r"\s+", "", (text or "")).lower()


def dedupe_receipt_matches(rows):
    """영수증·문자 캡처 행이 카드사 내역 행과 같은 거래면 병합해 이중 계상을 막는다
    (collect.md "하는 단계" 4-6, "판단 기준" 날짜+금액+결제처+결제수단 매칭).

    날짜·금액·결제수단이 같고 결제처도 같으면(정규화 후 정확 일치 — 일반 코드) 같은
    거래로 보고, 영수증의 구매항목만 카드 내역 행에 옮긴 뒤 영수증 행은 버린다(카드
    내역이 금액의 정본). 날짜·금액·결제수단은 같은데 결제처만 다르면 확신할 수 없어
    자동 병합하지 않고 두 행 모두 확인 필요로 낮춰 사람이 review하게 한다
    (§ 못 할 때 "영수증-카드내역 매칭이 불확실하면 반려 대신 review").
    """
    card_rows = [r for r in rows if r.get("source_type") == "card_excel"]
    receipt_rows = [r for r in rows if r.get("source_type") == "receipt"]
    other_rows = [r for r in rows if r.get("source_type") not in ("card_excel", "receipt")]
    if not card_rows or not receipt_rows:
        return rows

    def match_key(r):
        try:
            amount = round(float(r.get("금액") or 0), 2)
        except (TypeError, ValueError):
            amount = r.get("금액")
        return (r.get("날짜"), amount, _norm_for_match(r.get("결제수단")))

    card_by_key = {}
    for r in card_rows:
        card_by_key.setdefault(match_key(r), []).append(r)

    kept_receipts = []
    for rec in receipt_rows:
        candidates = card_by_key.get(match_key(rec), [])
        exact = [c for c in candidates if _norm_for_match(c["결제처"]) == _norm_for_match(rec["결제처"])]
        if exact:
            target = exact[0]
            if rec.get("구매항목") and not target.get("구매항목"):
                target["구매항목"] = rec["구매항목"]
            continue  # 카드 내역이 정본 — 영수증 행은 흡수하고 버린다 (이중 계상 방지)
        if candidates:  # 날짜·금액·결제수단은 같은데 결제처가 다름 — 불확실, review로 남긴다
            for c in candidates:
                c["collect_status"] = "확인 필요"
            rec["collect_status"] = "확인 필요"
            kept_receipts.append(rec)
        else:
            kept_receipts.append(rec)
    return other_rows + card_rows + kept_receipts


def map_table_row(r):
    """call_agent_convert_table 출력 행 → 거래 표 스키마 v1 (transaction_id는 나중에 부여)."""
    income = r.get("수익") or 0
    expense = r.get("지출") or 0
    amount = income if income > 0 else -expense
    payer_hint = f"{r.get('결제자') or ''} {r.get('결제수단') or ''}"
    foreign_amount = r.get("원거래금액") or 0
    return {
        "transaction_id": "",
        "날짜": r.get("날짜") or "",
        "금액": amount,
        "결제처": r.get("결제처") or "",
        "카테고리": "",           # 가공(refine) 몫
        "비고": r.get("비고") or "",
        "결제수단": r.get("결제수단") or "",
        "결제구분": "법인결제" if "법인" in payer_hint else "개인결제",
        # 해외결제면 AI 변환이 원본 통화·해외금액을 함께 내놓는다 — 국내 결제는 빈칸 유지
        "원거래통화": (r.get("원거래통화") or "").strip(),
        "원거래금액": foreign_amount if foreign_amount else "",
        "source_type": "card_excel",
        "collect_status": "확인됨" if r.get("확인됨") else "확인 필요",
        "구매항목": r.get("구매항목") or "",
    }


def map_receipt(r):
    """call_agent_with_image 출력 → 거래 표 스키마 v1 한 행."""
    amount = r.get("금액") or 0
    return {
        "transaction_id": "",
        "날짜": r.get("날짜") or "",
        "금액": -amount if amount > 0 else amount,  # 영수증 금액은 총 결제액 → 지출 음수
        "결제처": r.get("결제처") or "",
        "카테고리": "",
        "비고": "",
        "결제수단": r.get("결제수단") or "",
        "결제구분": "개인결제",   # 영수증만으로 법인 여부 판단 근거 없음 — 기본값
        "원거래통화": "",
        "원거래금액": "",
        "source_type": "receipt",
        "collect_status": "확인됨" if r.get("확인됨") else "확인 필요",
        "구매항목": r.get("구매항목") or "",
    }


def load_manual_entries():
    if not os.path.exists(MANUAL_ENTRIES_PATH):
        return []
    try:
        with open(MANUAL_ENTRIES_PATH, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return []


def map_manual_entry(m):
    """POST /uploads/manual로 받은 직접 입력 한 건 → 거래 표 스키마 v1 한 행.

    값을 사용자가 직접 입력했으므로 핵심 값(날짜·금액·결제처)을 못 읽는 경우가 없다 —
    collect_status는 항상 확인됨 (collect.md "하는 단계 7").
    """
    return {
        "transaction_id": "",
        "날짜": m.get("날짜") or "",
        "금액": m.get("금액") or 0,
        "결제처": m.get("결제처") or "",
        "카테고리": "",
        "비고": m.get("비고") or "",
        "결제수단": m.get("결제수단") or "현금",
        "결제구분": m.get("결제구분") or "개인결제",
        "원거래통화": "",
        "원거래금액": "",
        "source_type": "manual",
        "collect_status": "확인됨",
        "구매항목": "",
    }


def process_image(path):
    """이미지 1건 처리 — (rows, error) 반환. AI 실패는 error 문자열로 흡수한다.

    영수증 판독은 정확도를 위해 개별 호출을 유지한다 (collect.md "하는 단계 1").
    """
    name = os.path.basename(path)
    try:
        ext = os.path.splitext(name)[1].lower()
        data = call_agent.call_agent_with_image(open(path, "rb").read(), IMAGE_MEDIA_TYPES[ext])
        return [map_receipt(data)], None
    except Exception as e:
        return [], f"{name}: 처리 실패 — {e}"


def process_table_batch(paths):
    """낯선 서식 파일 여러 건을 한 번의 AI 호출로 변환 — (rows, errors) 반환.

    호출당 고정 비용(시스템 프롬프트·CLI 세션 오버헤드)을 줄이기 위해 묶는다
    (collect.md "하는 단계 1" 확정). 읽기 실패·빈 파일은 개별 오류로 흡수한다.
    """
    sections, errors = [], []
    for path in paths:
        name = os.path.basename(path)
        try:
            ext = os.path.splitext(name)[1].lower()
            if ext == ".xlsx":
                text = xlsx_to_text(path)
            elif ext == ".pdf":
                text = pdf_to_text(path)
            else:
                text = read_text_any(path)
        except Exception as e:
            errors.append(f"{name}: 처리 실패 — {e}")
            continue
        if not text.strip():
            errors.append(f"{name}: 내용이 비어 있음")
            continue
        sections.append(f"=== 파일: {name} ===\n{text}")
    if not sections:
        return [], errors
    combined = ("아래에 여러 파일의 표가 '=== 파일: 이름 ===' 구분선으로 이어진다. "
                "구분선은 데이터가 아니니 행으로 변환하지 않는다.\n\n" + "\n\n".join(sections))
    try:
        rows = call_agent.call_agent_convert_table(combined)
        return [map_table_row(r) for r in rows], errors
    except Exception as e:
        names = ", ".join(os.path.basename(p) for p in paths)
        return [], errors + [f"{names}: 일괄 변환 실패 — {e}"]


def assign_ids(rows, month):
    """하나카드 정규화가 이미 쓴 tx_{YYMMDD}_{NN} 순번을 이어서 나머지 행에 id를 부여한다.

    날짜를 못 읽은 행은 대상 월 기반 자리표 키(YYMM00)를 쓴다 — 행을 버리지 않는다.
    """
    day_seq = {}
    id_pat = re.compile(r"tx_(\d{6})_(\d+)")
    for r in rows:
        m = id_pat.fullmatch(r.get("transaction_id") or "")
        if m:
            day_seq[m.group(1)] = max(day_seq.get(m.group(1), 0), int(m.group(2)))
    for r in rows:
        if r.get("transaction_id"):
            continue
        m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", r.get("날짜") or "")
        key = f"{m.group(1)[2:]}{m.group(2)}{m.group(3)}" if m else month.replace("-", "")[2:] + "00"
        day_seq[key] = day_seq.get(key, 0) + 1
        r["transaction_id"] = f"tx_{key}_{day_seq[key]:02d}"


def build_envelope(rows, stats, errors):
    """interface-spec.md "단계 결과 보고" 규격 — 상태 미확인 행은 확인 필요, 파일 실패는 오류."""
    total = len(rows)
    message = (f"업로드 처리 — 카드사 파일 {stats['card']}건 · 이미지 {stats['image']}장"
               + (f" · 직접 입력 {stats['manual']}건" if stats.get("manual") else "")
               + (f" · 실패 {len(errors)}건" if errors else ""))
    if total == 0:
        return {"stage": "collect", "status": "empty", "output": "",
                "counts": {"total": 0, "ok": 0, "flagged": 0},
                "flags": [{"row": 0, "type": "오류", "reason": e} for e in errors],
                "message": message if errors else "수집 대상 없음"}
    row_flags = []
    for idx, r in enumerate(rows, start=1):
        if r["collect_status"] != "확인됨":
            core_missing = not r["날짜"] or not r["금액"] or not r["결제처"]
            row_flags.append({
                "row": idx,
                "type": "오류" if (r["source_type"] == "receipt" and core_missing) else "확인 필요",
                "reason": f"{r['결제처'] or '결제처 미상'} 거래 — 핵심 값 확인 필요",
            })
    # 파일 단위 실패는 flags에만 싣고 counts에는 넣지 않는다 (counts는 행 단위 — merge 관례와 동일)
    flags = [{"row": 0, "type": "오류", "reason": e} for e in errors] + row_flags
    return {
        "stage": "collect",
        "status": "ok" if not flags else "partial",
        "output": "collect/result.csv",
        "counts": {"total": total, "ok": total - len(row_flags), "flagged": len(row_flags)},
        "flags": flags,
        "message": message,
    }


def run(month, upload_dir=DEFAULT_UPLOAD_DIR, on_progress=None):
    """파이프라인 진입점 — collect.run()과 동일 계약.

    on_progress(done, total): 처리한 파일 수 기준 진행 통지 (선택 — 지휘의 웹 진행률 표시용,
    보고 규격·산출물에는 영향 없음).
    """
    paths = sorted(e.path for e in os.scandir(upload_dir) if e.is_file()) if os.path.isdir(upload_dir) else []
    manual_entries = load_manual_entries()
    if not paths and not manual_entries:
        return {"out_path": None, "rows": [], "envelope": build_envelope([], {"card": 0, "image": 0, "manual": 0}, [])}

    # zip은 압축을 풀어 안의 파일을 개별 원천으로 편입한다 (collect.md "하는 단계 0" 확정).
    # 풀린 파일은 임시 폴더에만 쓰고 처리가 끝나면 지운다 — uploads/inbox/에는 원본 zip만 남는다.
    zip_paths = [p for p in paths if os.path.splitext(p)[1].lower() == ".zip"]
    other_paths = [p for p in paths if p not in zip_paths]
    zip_errors = []
    tempdir = tempfile.mkdtemp(prefix="collect_zip_") if zip_paths else None
    try:
        for zp in zip_paths:
            extracted, errs = extract_zip(zp, tempdir)
            other_paths += extracted
            zip_errors += errs
        paths = sorted(other_paths)

        total_files = len(paths)
        done_files = 0

        def notify():
            if on_progress:
                try:
                    on_progress(done_files, total_files)
                except Exception:
                    pass  # 관찰자 오류가 수집을 멈추면 안 된다

        notify()

        hana_paths, shinhan_paths, image_paths, table_paths = [], [], [], []
        for path in paths:
            ext = os.path.splitext(path)[1].lower()
            if ext in IMAGE_MEDIA_TYPES:
                image_paths.append(path)
            elif ext in TEXT_EXTS and is_hana_csv(read_text_any(path)):
                hana_paths.append(path)
            elif ext in TEXT_EXTS and is_shinhan_csv(read_text_any(path)):
                shinhan_paths.append(path)
            else:
                table_paths.append(path)

        rows, errors = [], list(zip_errors)
        if hana_paths:
            by_month = collect.normalize(hana_paths, default_month=month)
            rows += [r for month_rows in by_month.values() for r in month_rows]
            done_files += len(hana_paths)
            notify()

        if shinhan_paths:
            rows += collect.parse_shinhan(shinhan_paths)  # transaction_id는 assign_ids가 이어서 부여
            done_files += len(shinhan_paths)
            notify()

        if table_paths:
            batch_rows, batch_errors = process_table_batch(table_paths)
            rows += batch_rows
            errors += batch_errors
            done_files += len(table_paths)
            notify()

        if image_paths:
            with ThreadPoolExecutor(max_workers=AI_WORKERS) as pool:
                for file_rows, error in pool.map(process_image, image_paths):
                    rows += file_rows
                    if error:
                        errors.append(error)
                    done_files += 1
                    notify()
    finally:
        if tempdir:
            shutil.rmtree(tempdir, ignore_errors=True)

    rows += [map_manual_entry(m) for m in manual_entries]  # 현금 등 직접 입력 (collect.md "하는 단계 7")
    rows = dedupe_receipt_matches(rows)  # 영수증-카드내역 같은 거래 병합 (이중 계상 방지)
    for r in rows:  # 통화 표기 승격 안전망 — 원천(정형·AI 변환·영수증) 공통, 채워진 값은 유지
        promote_foreign(r)
    assign_ids(rows, month)
    rows.sort(key=lambda r: (r["날짜"], r["transaction_id"]))

    out_path = os.path.join(BASE_DIR, "result.csv")
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=collect.OUT_FIELDS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    stats = {"card": len(hana_paths) + len(shinhan_paths) + len(table_paths),
             "image": len(image_paths), "manual": len(manual_entries)}
    return {"out_path": out_path if rows else None, "rows": rows,
            "envelope": build_envelope(rows, stats, errors)}


if __name__ == "__main__":
    if len(sys.argv) < 2 or not re.fullmatch(r"\d{4}-\d{2}", sys.argv[1]):
        print("사용법: python3 collect/collect_uploads.py <대상 월 YYYY-MM> [업로드 폴더]")
        sys.exit(1)
    result = run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else DEFAULT_UPLOAD_DIR)
    env = result["envelope"]
    print(f"{env['status']}: {env['counts']['total']}건 — {env['message']}")
