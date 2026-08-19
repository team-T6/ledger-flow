"""collect/screen.html과 call-agent.py를 이어주는 로컬 서버.

screen.html을 정적 파일로 내주고, POST /call-agent·POST /read-receipt·POST /convert-table
요청을 받으면 call-agent.py의 call_agent()·call_agent_with_image()·call_agent_convert_table()을
그대로 호출해 결과를 돌려준다. POST /read-excel은 진짜 바이너리 엑셀(.xls/.xlsx)을
받아 서버에서 표로 파싱해 CSV 텍스트로 돌려준다 — 브라우저는 엑셀 파서가 없어 이 화면이
대신 읽어준다. POST /save-result는 collect/result.csv(산출물)를 디스크에 써준다. 단계
결과 보고는 파일로 남기지 않는다 — 보고는 파이프라인 실행 시 수집 호출의 응답(반환값)
으로 지휘에게 직접 전달한다(interface-spec.md §단계 결과 보고 전달 방식 확정). 이
화면·서버는 데이터 준비 도구다.
"""

import base64
import csv
import datetime
import http.server
import importlib.util
import io
import json
import socketserver
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
PORT = 8787

_spec = importlib.util.spec_from_file_location("call_agent_module", BASE_DIR / "call-agent.py")
call_agent_module = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(call_agent_module)


def _cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime.datetime, datetime.date)):
        if isinstance(value, datetime.datetime) and (value.hour, value.minute, value.second) != (0, 0, 0):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    return str(value)


def _xls_to_rows(raw_bytes):
    """구형 바이너리 xls(OLE2) — xlrd로 첫 시트를 읽는다."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = wb.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(_cell_to_text(xlrd.xldate_as_datetime(cell.value, wb.datemode)))
            else:
                row.append(_cell_to_text(cell.value))
        rows.append(row)
    return rows


def _xlsx_to_rows(raw_bytes):
    """신형 xlsx(zip) — openpyxl로 활성 시트를 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    ws = wb.active
    return [[_cell_to_text(v) for v in row] for row in ws.iter_rows(values_only=True)]


def excel_bytes_to_csv(raw_bytes):
    """바이너리 엑셀 파일(.xls/.xlsx)을 첫 시트 기준 CSV 텍스트로 변환한다."""
    if raw_bytes[:2] == b"PK":
        rows = _xlsx_to_rows(raw_bytes)
    elif raw_bytes[:4] == b"\xd0\xcf\x11\xe0":
        rows = _xls_to_rows(raw_bytes)
    else:
        raise ValueError("알 수 없는 파일 형식 (xls/xlsx 시그니처 아님)")

    out = io.StringIO()
    writer = csv.writer(out)
    for row in rows:
        writer.writerow(row)
    return out.getvalue()


class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def do_POST(self):
        if self.path == "/call-agent":
            self._handle_call_agent()
        elif self.path == "/read-receipt":
            self._handle_read_receipt()
        elif self.path == "/convert-table":
            self._handle_convert_table()
        elif self.path == "/read-excel":
            self._handle_read_excel()
        elif self.path == "/save-result":
            self._handle_save_result()
        else:
            self.send_error(404)

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(length) or b"{}")

    def _handle_call_agent(self):
        body = self._read_json_body()
        text = body.get("text", "")

        try:
            result = call_agent_module.call_agent(text)
            payload = {"result": result}
            status = 200
        except Exception as e:
            payload = {"error": str(e)}
            status = 500

        self._send_json(status, payload)

    def _handle_read_receipt(self):
        body = self._read_json_body()
        image_b64 = body.get("image", "")
        media_type = body.get("media_type", "")

        try:
            image_bytes = base64.b64decode(image_b64) if image_b64 else b""
            result = call_agent_module.call_agent_with_image(image_bytes, media_type)
            payload = {"result": result}
            status = 200
        except Exception as e:
            payload = {"error": str(e)}
            status = 500

        self._send_json(status, payload)

    def _handle_convert_table(self):
        body = self._read_json_body()
        text = body.get("text", "")

        try:
            result = call_agent_module.call_agent_convert_table(text)
            payload = {"result": result}
            status = 200
        except Exception as e:
            payload = {"error": str(e)}
            status = 500

        self._send_json(status, payload)

    def _handle_read_excel(self):
        body = self._read_json_body()
        data_b64 = body.get("data", "")

        try:
            raw_bytes = base64.b64decode(data_b64) if data_b64 else b""
            csv_text = excel_bytes_to_csv(raw_bytes)
            payload = {"result": {"csv": csv_text}}
            status = 200
        except Exception as e:
            payload = {"error": str(e)}
            status = 500

        self._send_json(status, payload)

    def _handle_save_result(self):
        body = self._read_json_body()
        csv_text = body.get("csv", "")

        try:
            (BASE_DIR / "result.csv").write_text(csv_text, encoding="utf-8")
            payload = {"ok": True}
            status = 200
        except Exception as e:
            payload = {"error": str(e)}
            status = 500

        self._send_json(status, payload)

    def _send_json(self, status, payload):
        response = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(response)))
        self.end_headers()
        self.wfile.write(response)


if __name__ == "__main__":
    with socketserver.TCPServer(("127.0.0.1", PORT), Handler) as httpd:
        print(f"http://127.0.0.1:{PORT}/screen.html 에서 확인하세요")
        httpd.serve_forever()
