"""merge 최종 산출물(result.xlsx, result.pdf)을 만드는 스크립트.

입력은 앞 단계의 실제 산출물 — 가공 거래 표 refine/result.csv(거래 표 스키마 v1)와
두 검증 결과 verify1/result.csv·verify2/result.csv(입력 행 + verify{n}_result·verify{n}_reason,
interface-spec.md 분류/기간·금액 검증 행 확정). 행 대조 키는 transaction_id.
가공 산출물이 없으면 취합 불가라 failed로 보고하고, 검증 결과가 없으면 그 자리는
"미완" 표시 + 사유를 남기고 나머지를 정상 진행한다 (단계 문서 "못 할 때").

한글 PDF 폰트는 merge/fonts/NanumGothic-Regular.ttf(OFL 라이선스, 리포에 포함)를 쓴다 —
운영체제마다 다른 시스템 폰트 경로에 기대지 않기 위함이다.

사용법: python3 merge/build_result.py
"""

import csv
import datetime
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "result.xlsx")
PDF_PATH = os.path.join(BASE_DIR, "result.pdf")
FONT_PATH = os.path.join(BASE_DIR, "fonts", "NanumGothic-Regular.ttf")
FONT_NAME = "NanumGothic"

# 지휘에게 보내는 단계 결과 보고의 output 필드 — repo 기준 상대 경로 (interface-spec.md 예시와 동일 관례)
OUTPUT_PATHS = ["merge/result.xlsx", "merge/result.pdf"]

REPO_ROOT = os.path.dirname(BASE_DIR)
REFINE_CSV = os.path.join(REPO_ROOT, "refine", "result.csv")
# 지휘의 월별 보관함 — 전월 대비·월별 추이의 데이터 원천 (summary.json만 읽기 전용으로 접근)
ARCHIVE_DIR = os.path.join(REPO_ROOT, "archive")
VERIFY_CSVS = {
    "verify1": os.path.join(REPO_ROOT, "verify1", "result.csv"),
    "verify2": os.path.join(REPO_ROOT, "verify2", "result.csv"),
}
# 부정 사용 검증(법인카드 부정 사용 감지)은 토글 옵트인 — 파일이 없으면 "미완"이 아니라 정상 생략
# (interface-spec.md 부정 사용 검증 행 · 2026-08-31 확정 로그)
VERIFY3_CSV = os.path.join(REPO_ROOT, "verify3", "result.csv")

# 엑셀 회계장부 컬럼 — 거래 표 스키마 (확정 v1)와 동일 (interface-spec.md "산출물 양식")
LEDGER_COLUMNS = ["transaction_id", "날짜", "금액", "결제처", "카테고리", "비고",
                  "결제수단", "결제구분", "원거래통화", "원거래금액",
                  "source_type", "source_file", "collect_status", "구매항목"]

# 데모(web/index.html) "장부 에디토리얼" 테마와 톤을 맞춘 팔레트
PAGE_BG = colors.HexColor("#FAF7F2")
INK = colors.HexColor("#1C1917")
MUTED = colors.HexColor("#7A7263")
ACCENT = colors.HexColor("#1E5B45")  # 제목·섹션 띠·표 헤더 — 구조적 요소
ACCENT_CONTRAST = colors.HexColor("#FAF7F2")
ACCENT_LIGHT = colors.HexColor("#EFE9DD")
CHART_COLOR = ACCENT  # 단일 시리즈 막대그래프는 accent 한 가지 색만 쓴다
# 파이 조각용 시리즈 팔레트 — accent 녹색 계열을 밝기로 벌리고, 조각이 많으면 보조 톤으로 이어간다
SERIES_COLORS = [colors.HexColor(c) for c in (
    "#1E5B45", "#3A7D63", "#5C9C82", "#8FBCA6", "#C2D9CB",
    "#A63A2E", "#C97B4A", "#D9B36A", "#7A7263", "#B5AC9A",
)]
FLAG_COLOR = colors.HexColor("#A63A2E")
FLAG_LIGHT = colors.HexColor("#F3E3DF")
GRID_COLOR = colors.HexColor("#E6DFD2")

# 표·그래프·섹션 배너가 전부 같은 좌우 여백을 쓰게 만드는 기준값 —
# 이 폭에서 벗어나는 요소가 없어야 열(오른쪽 끝)이 서로 어긋나지 않는다.
PAGE_MARGIN = 18 * mm
CONTENT_WIDTH = A4[0] - 2 * PAGE_MARGIN


def _col_widths(*fractions, total=CONTENT_WIDTH):
    """비율을 표 컬럼 폭(포인트)으로 바꾼다. 마지막 컬럼이 남은 폭을 다 가져가
    반올림 오차 없이 합이 정확히 total(기본 CONTENT_WIDTH)이 되게 한다."""
    widths = [total * f for f in fractions[:-1]]
    widths.append(total - sum(widths))
    return widths

def _parse_amount(value):
    """금액 문자열을 정수로 바꾼다 (콤마 허용). 해석 불가·빈 값은 None."""
    s = str(value or "").replace(",", "").strip()
    try:
        return int(s)
    except ValueError:
        return None


def load_transactions(fraud_check=False):
    """가공 거래 표(스키마 v1)에 검증 판정을 transaction_id로 붙여 돌려준다.

    fraud_check: 이번 실행의 부정 사용 감지 토글 값 — 켠 실행인데 verify3/result.csv가
    없으면(검증 실패) 토글을 끈 실행과 구분해 "미완"으로 남긴다.
    반환: (rows, incomplete) — incomplete는 검증 결과가 통째로 없는 쪽의 "미완" 사유 목록.
    화면(screen.html)이 아직 쓰는 구 필드(지출/수익/결제자)는 금액·결제구분에서 파생해 채운다.
    """
    with open(REFINE_CSV, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    incomplete = []
    verdicts = {}  # stage -> {transaction_id: (result, reason)}
    for stage, path in VERIFY_CSVS.items():
        if not os.path.exists(path):
            incomplete.append(f"{stage} 결과 없음 ({os.path.relpath(path, REPO_ROOT)}) — 미완 처리")
            verdicts[stage] = None
            continue
        with open(path, encoding="utf-8-sig", newline="") as f:
            verdicts[stage] = {
                v["transaction_id"]: (v.get(f"{stage}_result", ""), v.get(f"{stage}_reason", ""))
                for v in csv.DictReader(f)
            }

    # 부정 사용 검증은 토글을 켠 실행에만 존재 — 파일이 있으면 같은 방식으로 판정을 붙인다.
    # 토글을 켠 실행인데 파일이 없으면(검증 실패) 토글 끈 실행과 구분해 "미완"으로 남긴다
    # (orchestrator.md "판단 규칙" — 검증 편측 실패는 통합에 미완으로 전달, verify1·2와 동일 취급)
    verify3 = None
    if os.path.exists(VERIFY3_CSV):
        with open(VERIFY3_CSV, encoding="utf-8-sig", newline="") as f:
            verify3 = {
                v["transaction_id"]: (v.get("verify3_result", ""), v.get("verify3_reason", ""))
                for v in csv.DictReader(f)
            }
    elif fraud_check:
        incomplete.append(f"verify3 결과 없음 ({os.path.relpath(VERIFY3_CSV, REPO_ROOT)}) — 미완 처리")

    for row in rows:
        amount = _parse_amount(row.get("금액"))
        # 화면 호환용 파생 필드 — 스키마 v1의 금액(부호)·결제구분에서 계산
        row["지출"] = -amount if (amount is not None and amount < 0) else None
        row["수익"] = amount if (amount is not None and amount > 0) else None
        row["결제자"] = row.get("결제구분", "")
        for stage in VERIFY_CSVS:
            if verdicts[stage] is None:
                row[f"{stage}_result"], row[f"{stage}_reason"] = "미완", f"{stage} 결과 없음"
            else:
                result, reason = verdicts[stage].get(row.get("transaction_id"), ("미완", f"{stage} 결과에 해당 행 없음"))
                row[f"{stage}_result"], row[f"{stage}_reason"] = result, reason
        if verify3 is not None:
            result, reason = verify3.get(row.get("transaction_id"), ("미완", "verify3 결과에 해당 행 없음"))
            row["verify3_result"], row["verify3_reason"] = result, reason
    return rows, incomplete


def split_transactions(transactions):
    """verify1·verify2 중 하나라도 반려면 확인 필요 목록으로 뺀다 (재시도 없음).
    부정 사용 검증(있으면)의 확인 요청 건은 **장부·집계에서 빼지 않고** 확인 필요 목록에만
    함께 싣는다 — 데이터는 정상이고 업무 사용 여부만 확인 대상이라서다
    (단계 문서 merge.md · interface-spec.md 부정 사용 검증 확정 로그). 그래서 이 건은
    ok_rows와 flagged_rows 양쪽에 들어간다 — 보고 counts.ok는 build_envelope가
    total - flagged로 계산해 문제 건으로 센다.
    flagged_rows에는 지휘 보고의 flags[].row로 쓸 1-기준 행 번호(row)를 함께 담는다."""
    ok_rows, flagged_rows = [], []
    for idx, row in enumerate(transactions, start=1):
        if row["verify1_result"] == "반려":
            flagged_rows.append({**row, "row": idx, "type": "반려",
                                 "reason": f"{row.get('verify1_reason', '')} (verify1)"})
        elif row["verify2_result"] == "반려":
            flagged_rows.append({**row, "row": idx, "type": "반려",
                                 "reason": f"{row.get('verify2_reason', '')} (verify2)"})
        else:
            if row.get("verify3_result") == "확인 요청":
                flagged_rows.append({**row, "row": idx, "type": "확인 요청",
                                     "reason": f"{row.get('verify3_reason', '')} (verify3)"})
            ok_rows.append(row)
    return ok_rows, flagged_rows


def build_row_list(transactions, flagged_rows):
    """결과·보관 화면의 "거래 내역" 표용 — 전체 거래를 확인 필요 여부·사유와 함께 반환한다.
    split_transactions()가 매긴 1-기준 행 번호(row)로 flagged_rows와 대조한다."""
    reason_by_row = {r["row"]: r["reason"] for r in flagged_rows}
    rows = []
    for idx, r in enumerate(transactions, start=1):
        rows.append({
            "transaction_id": r.get("transaction_id", ""),
            "날짜": r.get("날짜", ""),
            "결제처": r.get("결제처", ""),
            "카테고리": r.get("카테고리", ""),
            "지출": r.get("지출") or 0,
            "수익": r.get("수익") or 0,
            "결제수단": r.get("결제수단", ""),
            "결제구분": r.get("결제구분", ""),
            "flagged": idx in reason_by_row,
            "reason": reason_by_row.get(idx, ""),
        })
    return rows


def build_envelope(total, ok_rows, flagged_rows, failed=False, message="", incomplete=()):
    """지휘(orchestrator)에게 돌려주는 단계 결과 보고 — interface-spec.md "단계 결과 보고" 규격.
    status 어휘 4개 고정: ok(정상) · empty(대상 없음) · partial(확인 필요 건 있음) · failed(단계 실패).
    output은 empty·failed일 때 빈 값(interface-spec.md "단계 결과 보고" 필드 설명).
    incomplete는 검증 결과 부재 등 자리 단위 "미완" 사유 목록 — flags에 type 미완으로 싣는다."""
    if failed:
        status, output = "failed", []
    elif total == 0:
        status, output = "empty", []
    else:
        status = "ok" if not (flagged_rows or incomplete) else "partial"
        output = OUTPUT_PATHS

    flags = [{"row": 0, "type": "미완", "reason": reason} for reason in incomplete]
    # 부정 사용 검증의 확인 요청 건은 유형을 보존한다 (장부에는 실리고 확인 목록에만 표시)
    flags += [{"row": r["row"],
               "type": "확인 요청" if r.get("type") == "확인 요청" else "확인 필요",
               "reason": r["reason"]} for r in flagged_rows]
    return {
        "stage": "merge",
        "status": status,
        "output": output,
        # 확인 요청 건이 ok_rows(장부)와 flagged_rows 양쪽에 들어가므로
        # ok는 len(ok_rows)가 아니라 total - flagged로 센다 (문제 건 제외)
        "counts": {"total": total, "ok": max(0, total - len(flagged_rows)),
                   "flagged": len(flagged_rows)},
        "flags": flags,
        "message": message,
    }


def _cell_width(value):
    """열 너비 산정용 표시 폭 — 한글 등 전각 문자는 반각의 2배로 센다."""
    text = "" if value is None else str(value)
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in text)


def build_ledger(ok_rows, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "회계장부"
    ws.append(LEDGER_COLUMNS)
    for row in ok_rows:
        ws.append([row.get(col) for col in LEDGER_COLUMNS])

    # 스타일 — docs/design-guide.md §5: 헤더 = accent 채움 + 종이색 굵은 글자,
    # 열 너비 = 내용 길이에 맞춰 사전 지정 (열면 바로 읽히게, 수동 조정 불필요)
    header_fill = PatternFill("solid", start_color="1E5B45")
    header_font = Font(bold=True, color="FAF7F2")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for idx, col in enumerate(LEDGER_COLUMNS, start=1):
        widest = max(
            [_cell_width(col)] + [_cell_width(row.get(col)) for row in ok_rows]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(widest + 4, 44)

    wb.save(xlsx_path)


def _styles():
    return {
        "title": ParagraphStyle("title", fontName=FONT_NAME, fontSize=24, leading=28, textColor=ACCENT),
        "meta": ParagraphStyle("meta", fontName=FONT_NAME, fontSize=10, leading=14, textColor=MUTED),
        "body": ParagraphStyle("body", fontName=FONT_NAME, fontSize=11, leading=16, textColor=INK),
        "flag": ParagraphStyle("flag", fontName=FONT_NAME, fontSize=10, leading=14, textColor=FLAG_COLOR),
        # 표 셀 안에서 줄바꿈이 필요한 일반 텍스트(긴 결제처명 등)용
        "cell": ParagraphStyle("cell", fontName=FONT_NAME, fontSize=10, leading=13, textColor=INK),
        # 인사이트 요약문 불릿용
        "insight": ParagraphStyle("insight", fontName=FONT_NAME, fontSize=10, leading=15, textColor=INK),
    }


def _fmt_delta(cur, before):
    """전월 대비 증감 표기 — "+210,000원 (+7.2%)". 기준값 0이면 %는 생략한다."""
    diff = cur - before
    pct = f" ({diff / before * 100:+.1f}%)" if before else ""
    return f"{diff:+,}원{pct}"


def _is_weekend(date_str):
    """YYYY-MM-DD가 토·일이면 True. 날짜 해석 불가면 False (인사이트는 부가 정보)."""
    try:
        return datetime.date.fromisoformat(str(date_str or "").strip()).weekday() >= 5
    except ValueError:
        return False


def _group_top(items, limit=9):
    """(이름, 금액) 내림차순 목록의 상위 limit개만 남기고 나머지를 한 조각으로 묶는다 —
    파이 조각이 잘게 쪼개지는 것 방지. 묶음 이름은 실제 카테고리 "기타"와 헷갈리지 않게
    "그 외 N개"로 쓴다. 묶어서 조각이 하나만 줄면(초과분 1개) 묶지 않고 그대로 보여준다."""
    if len(items) <= limit + 1:
        return list(items)
    rest = sum(amount for _, amount in items[limit:])
    return list(items[:limit]) + [(f"그 외 {len(items) - limit}개", rest)]


def _insight_box(lines, style):
    """인사이트 요약문 묶음 — 옅은 배경 상자에 불릿 줄로 싣는다."""
    rows = [[Paragraph(f"•  {line}", style)] for line in lines]
    return Table(rows, colWidths=[CONTENT_WIDTH], style=TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), ACCENT_LIGHT),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (0, 0), 6),
        ("BOTTOMPADDING", (0, -1), (0, -1), 6),
    ]))


def _section_heading(text):
    """섹션 제목 띠 — 아래 표(_table)와 똑같이 colWidths=[CONTENT_WIDTH]인 표로 만든다.
    Paragraph의 backColor는 leftIndent/borderPadding에 따라 표와 다른 폭으로 그려질 수 있어,
    표와 좌우 끝을 정확히 맞추려고 같은 Table 메커니즘을 그대로 재사용한다."""
    table = Table([[text]], colWidths=[CONTENT_WIDTH])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("BACKGROUND", (0, 0), (-1, -1), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, -1), ACCENT_CONTRAST),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def _table(header, rows, col_widths, num_cols=(), font_size=10):
    """헤더 색·격자선이 있는 표 하나를 만든다. num_cols는 오른쪽 정렬할 열 인덱스.
    font_size는 좁은 폭(그래프 옆 배치 등)에 맞춰 줄일 때 쓴다."""
    data = [header] + rows
    style = [
        ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), ACCENT_CONTRAST),
        ("GRID", (0, 0), (-1, -1), 0.5, GRID_COLOR),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PAGE_BG, ACCENT_LIGHT]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]
    for col in num_cols:
        style.append(("ALIGN", (col, 0), (col, -1), "RIGHT"))
    return Table(data, colWidths=col_widths, style=TableStyle(style))


def _pie_chart(items):
    """(이름, 금액) 목록을 파이 그래프 + 범례로 그린다 — 카테고리별 지출용.

    파이는 왼쪽, 범례는 오른쪽에 두고 전체 폭은 표·배너와 같은 CONTENT_WIDTH를 쓴다.
    조각 색은 SERIES_COLORS를 순서대로 쓰고, 조각이 팔레트보다 많으면 순환한다.
    """
    height = 68 * mm
    total = sum(amount for _, amount in items) or 1
    drawing = Drawing(CONTENT_WIDTH, height)

    pie = Pie()
    pie.x = 8 * mm
    pie.y = 6 * mm
    pie.width = pie.height = height - 12 * mm
    pie.data = [amount for _, amount in items]
    pie.labels = None
    pie.slices.strokeColor = PAGE_BG
    pie.slices.strokeWidth = 0.8
    for i in range(len(items)):
        pie.slices[i].fillColor = SERIES_COLORS[i % len(SERIES_COLORS)]
    drawing.add(pie)

    legend = Legend()
    legend.x = pie.x + pie.width + 14 * mm
    legend.y = height / 2
    legend.boxAnchor = "w"
    legend.fontName = FONT_NAME
    legend.fontSize = 9
    legend.columnMaximum = 10
    legend.alignment = "right"
    legend.colorNamePairs = [
        (SERIES_COLORS[i % len(SERIES_COLORS)],
         f"{name}  {amount:,}원 ({amount / total * 100:.1f}%)")
        for i, (name, amount) in enumerate(items)
    ]
    drawing.add(legend)
    return drawing


def _hbar_chart(items, total=None):
    """(이름, 금액) 목록을 가로 막대로 그린다 — 순위형(Top 10)·소수 항목 비교용.

    첫 항목이 맨 위에 오도록 뒤집어 넣는다 (HorizontalBarChart는 첫 카테고리를
    맨 아래에 그린다). 금액은 막대 끝 라벨로 붙이고 값 축은 숨긴다.
    total을 주면 라벨에 비중(%)도 붙는다 — 같은 수치의 표를 생략할 때 쓴다.
    """
    LEFT = 40 * mm   # 이름 라벨 자리
    RIGHT = 34 * mm if total else 24 * mm  # 막대 끝 라벨 자리 (비중 포함 시 더 넓게)
    TOP = BOTTOM = 2 * mm
    row = 8 * mm
    height = TOP + BOTTOM + row * len(items)

    drawing = Drawing(CONTENT_WIDTH, height)
    chart = HorizontalBarChart()
    chart.x = LEFT
    chart.y = BOTTOM
    chart.width = CONTENT_WIDTH - LEFT - RIGHT
    chart.height = height - TOP - BOTTOM

    ordered = list(reversed(items))
    chart.data = [[amount for _, amount in ordered]]
    chart.categoryAxis.categoryNames = [name for name, _ in ordered]
    chart.categoryAxis.labels.fontName = FONT_NAME
    chart.categoryAxis.labels.fontSize = 9
    chart.valueAxis.visible = 0
    chart.valueAxis.valueMin = 0

    chart.bars[0].fillColor = CHART_COLOR
    chart.bars[0].strokeColor = PAGE_BG
    chart.bars[0].strokeWidth = 0.4
    chart.barLabels.fontName = FONT_NAME
    chart.barLabels.fontSize = 9
    chart.barLabels.fillColor = INK
    chart.barLabels.boxAnchor = "w"
    chart.barLabels.nudge = 8
    if total:
        chart.barLabelFormat = lambda v: f"{int(v):,} ({v / total * 100:.1f}%)"
    else:
        chart.barLabelFormat = lambda v: f"{int(v):,}"

    drawing.add(chart)
    return drawing


def _share_band(items):
    """(이름, 금액) 목록을 100% 누적 가로 막대 하나로 그린다 — 개인/법인 구성비용.

    조각이 넓으면 라벨을 조각 안(흰 글씨), 좁으면 막대 아래(잉크색)에 둔다.
    금액 0인 항목은 그리지 않는다.
    """
    total = sum(amount for _, amount in items) or 1
    band_h = 9 * mm
    height = band_h + 6 * mm  # 아래 라벨 줄 여유
    drawing = Drawing(CONTENT_WIDTH, height)

    x = 0.0
    for i, (name, amount) in enumerate(items):
        if amount <= 0:
            continue
        w = CONTENT_WIDTH * amount / total
        drawing.add(Rect(x, height - band_h, w, band_h,
                         fillColor=SERIES_COLORS[i % len(SERIES_COLORS)],
                         strokeColor=PAGE_BG, strokeWidth=0.8))
        label = f"{name}  {amount:,}원 ({amount / total * 100:.1f}%)"
        if w >= 45 * mm:
            s = String(x + w / 2, height - band_h / 2 - 3.2, label, textAnchor="middle")
            s.fillColor = ACCENT_CONTRAST
        else:
            s = String(min(x, CONTENT_WIDTH - 45 * mm), 1, label, textAnchor="start")
            s.fillColor = INK
        s.fontName = FONT_NAME
        s.fontSize = 9
        drawing.add(s)
        x += w
    return drawing


def _trend_line_chart(trend):
    """월별 지출·수익 꺾은선 — 보관 월이 쌓여 5개 이상이면 막대 대신 쓴다.

    trend: (월, 총지출, 총수익) 목록 (월 오름차순).
    """
    LEFT_AXIS_MARGIN = 16 * mm
    BOTTOM_AXIS_MARGIN = 12 * mm
    RIGHT_PAD = 4 * mm
    TOP_PAD = 8 * mm
    height = 60 * mm

    drawing = Drawing(CONTENT_WIDTH, height)
    chart = HorizontalLineChart()
    chart.x = LEFT_AXIS_MARGIN
    chart.y = BOTTOM_AXIS_MARGIN
    chart.width = CONTENT_WIDTH - LEFT_AXIS_MARGIN - RIGHT_PAD
    chart.height = height - BOTTOM_AXIS_MARGIN - TOP_PAD

    chart.data = [[exp for _, exp, _ in trend], [inc for _, _, inc in trend]]
    chart.categoryAxis.categoryNames = [m for m, _, _ in trend]
    chart.categoryAxis.labels.fontName = FONT_NAME
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.labels.fontName = FONT_NAME
    chart.valueAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labelTextFormat = lambda v: f"{int(v):,}"
    chart.lines[0].strokeColor = ACCENT
    chart.lines[1].strokeColor = SERIES_COLORS[6]  # 보조 톤 — 수익
    chart.lines[0].strokeWidth = 1.6
    chart.lines[1].strokeWidth = 1.6
    drawing.add(chart)

    legend = Legend()
    legend.x = CONTENT_WIDTH - RIGHT_PAD
    legend.y = height - 3 * mm
    legend.boxAnchor = "ne"
    legend.fontName = FONT_NAME
    legend.fontSize = 8
    legend.alignment = "right"
    legend.colorNamePairs = [(ACCENT, "총지출"), (SERIES_COLORS[6], "총수익")]
    drawing.add(legend)
    return drawing


def load_archived_months():
    """지휘가 보관한 월별 집계(archive/<YYYY-MM>/summary.json)를 전부 읽는다.

    전월 대비·월별 추이는 부가 정보라, 보관함이 없거나 깨진 월은 조용히 건너뛰고
    리포트 생성을 실패시키지 않는다."""
    months = {}
    if not os.path.isdir(ARCHIVE_DIR):
        return months
    for name in os.listdir(ARCHIVE_DIR):
        if not re.fullmatch(r"\d{4}-\d{2}", name):
            continue
        path = os.path.join(ARCHIVE_DIR, name, "summary.json")
        try:
            with open(path, encoding="utf-8") as f:
                months[name] = json.load(f)
        except (OSError, ValueError):
            continue
    return months


def _prev_month(month):
    """YYYY-MM의 직전 달 — 연 경계(1월→전년 12월)도 넘어간다."""
    first = datetime.date(int(month[:4]), int(month[5:7]), 1)
    last_of_prev = first - datetime.timedelta(days=1)
    return f"{last_of_prev.year:04d}-{last_of_prev.month:02d}"


def _category_chart(by_category, width=CONTENT_WIDTH):
    """카테고리별 지출 막대그래프 — 도형(Drawing)으로 그린다.

    Drawing 폭은 기본으로 표·섹션 배너와 같은 CONTENT_WIDTH를 써서 오른쪽 끝이
    서로 어긋나지 않게 하고(표 옆에 나란히 둘 때는 width로 줄인다), 막대 사이
    간격(barSpacing)을 고정값으로 줘서 항상 일정하게 유지한다.
    """
    LEFT_AXIS_MARGIN = 16 * mm   # 금액 축 눈금 숫자가 들어갈 자리
    BOTTOM_AXIS_MARGIN = 14 * mm  # 카테고리 이름이 들어갈 자리
    RIGHT_PAD = 4 * mm
    TOP_PAD = 6 * mm
    height = 60 * mm

    drawing = Drawing(width, height)
    chart = VerticalBarChart()
    chart.x = LEFT_AXIS_MARGIN
    chart.y = BOTTOM_AXIS_MARGIN
    # 항목이 적으면 그래프 폭도 항목 수에 비례해 줄인다 — 막대가 전체 폭으로
    # 늘어지거나 라벨 위치와 어긋나는 것 방지 (항목당 28mm)
    max_plot_width = width - LEFT_AXIS_MARGIN - RIGHT_PAD
    chart.width = min(max_plot_width, len(by_category) * 28 * mm)
    chart.height = height - BOTTOM_AXIS_MARGIN - TOP_PAD

    chart.data = [[amount for _, amount in by_category]]
    chart.categoryAxis.categoryNames = [name for name, _ in by_category]
    chart.categoryAxis.labels.fontName = FONT_NAME
    chart.categoryAxis.labels.fontSize = 8

    chart.valueAxis.labels.fontName = FONT_NAME
    chart.valueAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labelTextFormat = lambda v: f"{int(v):,}"
    chart.valueAxis.visibleGrid = 1
    chart.valueAxis.gridStrokeColor = GRID_COLOR
    chart.valueAxis.gridStrokeWidth = 0.4

    chart.bars[0].fillColor = CHART_COLOR
    chart.bars[0].strokeColor = PAGE_BG
    chart.bars[0].strokeWidth = 0.4

    drawing.add(chart)
    return drawing


def summarize(ok_rows):
    """PDF 작성과 화면 미리보기가 함께 쓰는 집계 — 카테고리·결제수단·결제자별 합계, Top 지출.

    결제수단·결제구분이 빈 행은 "(미상)"으로 묶는다 (빈 이름이 표에 그대로 실리는 것 방지).
    구분별 카테고리 세부(by_payer_category)는 이원화 체계라 구분마다 자기 카테고리로 집계한다.
    """
    total_expense = sum(r["지출"] or 0 for r in ok_rows)
    total_income = sum(r["수익"] or 0 for r in ok_rows)

    # 카테고리·결제수단·결제구분 집계는 전부 "지출" 집계 — 수익 행(지출 None)은 제외한다
    # (수익 카테고리가 0원으로 지출 표에 끼는 것 방지)
    expense_rows = [r for r in ok_rows if r["지출"]]

    by_category = {}
    for r in expense_rows:
        by_category[r["카테고리"]] = by_category.get(r["카테고리"], 0) + r["지출"]

    by_method = {}
    for r in expense_rows:
        method = (r["결제수단"] or "").strip() or "(미상)"
        by_method[method] = by_method.get(method, 0) + r["지출"]

    by_payer = {}
    by_payer_category = {}
    for r in expense_rows:
        payer = (r["결제자"] or "").strip() or "(미상)"
        by_payer[payer] = by_payer.get(payer, 0) + r["지출"]
        by_payer_category.setdefault(payer, {})
        by_payer_category[payer][r["카테고리"]] = \
            by_payer_category[payer].get(r["카테고리"], 0) + r["지출"]

    top_spenders = sorted(expense_rows, key=lambda r: r["지출"], reverse=True)[:10]

    return {
        "total_expense": total_expense,
        "total_income": total_income,
        "by_category": sorted(by_category.items(), key=lambda kv: kv[1], reverse=True),
        "by_method": list(by_method.items()),
        "by_payer": list(by_payer.items()),
        "by_payer_category": {
            payer: sorted(cats.items(), key=lambda kv: kv[1], reverse=True)
            for payer, cats in by_payer_category.items()
        },
        "top_spenders": top_spenders,
    }


def _paint_page_background(canvas, doc):
    """데모 랜딩 페이지와 같은 크림색(PAGE_BG) 바탕지를 매 페이지에 깔고,
    페이지 번호를 오른쪽 아래에 찍는다."""
    canvas.saveState()
    canvas.setFillColor(PAGE_BG)
    canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
    canvas.setFont(FONT_NAME, 9)
    canvas.setFillColor(MUTED)
    canvas.drawRightString(A4[0] - PAGE_MARGIN, PAGE_MARGIN / 2, str(canvas.getPageNumber()))
    canvas.restoreState()


def _derive_month(rows):
    """대상 월 미전달 시 데이터에서 파생 — 날짜(YYYY-MM-DD)의 최빈 YYYY-MM."""
    counts = {}
    for r in rows:
        prefix = str(r.get("날짜") or "")[:7]
        if len(prefix) == 7:
            counts[prefix] = counts.get(prefix, 0) + 1
    return max(counts, key=counts.get) if counts else None


def build_report(ok_rows, flagged_rows, summary, pdf_path, month=None, incomplete=()):
    pdfmetrics.registerFont(TTFont(FONT_NAME, FONT_PATH))
    s = _styles()

    total_expense = summary["total_expense"]
    total_income = summary["total_income"]
    net = total_income - total_expense

    # 대상 월은 지휘가 전달한 실행 파라미터가 정본 — 없으면(단독 실행) 데이터에서 파생 표기
    base_month = month or _derive_month(ok_rows + flagged_rows)
    title_text = (f"{int(base_month[:4])}년 {int(base_month[5:7])}월 결산 리포트"
                  if base_month else "결산 리포트")
    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        topMargin=PAGE_MARGIN, bottomMargin=PAGE_MARGIN, leftMargin=PAGE_MARGIN, rightMargin=PAGE_MARGIN,
    )
    story = []

    story.append(Paragraph(title_text, s["title"]))
    # 대상 월은 제목에 이미 있어 부제에는 생성 일시만 둔다 (파생 표기면 추정임을 덧붙인다)
    subtitle = f"생성 {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if not month and base_month:
        subtitle += " · 대상 월은 데이터 기준 추정"
    story.append(Paragraph(subtitle, s["meta"]))
    story.append(Spacer(1, 8 * mm))

    story.append(_section_heading("1. 결산 개요"))
    story.append(Spacer(1, 2 * mm))
    # 카드(결제수단×결제구분)별 한 행 + 합계 행 — 장부에 실린 행(ok_rows) 기준
    by_card = {}
    for r in ok_rows:
        card = f"{(r.get('결제수단') or '').strip() or '(미상)'}({(r.get('결제구분') or '').strip() or '미상'})"
        entry = by_card.setdefault(card, {"expense": 0, "income": 0})
        entry["expense"] += r["지출"] or 0
        entry["income"] += r["수익"] or 0
    card_rows = [
        [card, f"{v['expense']:,}원", f"{v['income']:,}원", f"{v['income'] - v['expense']:,}원"]
        for card, v in sorted(by_card.items(), key=lambda kv: kv[1]["expense"], reverse=True)
    ]
    overview = _table(
        ["대상 카드", "총지출", "총수익", "순액"],
        card_rows + [["합계", f"{total_expense:,}원", f"{total_income:,}원", f"{net:,}원"]],
        col_widths=_col_widths(0.37, 0.21, 0.21, 0.21), num_cols=(1, 2, 3),
    )
    overview.setStyle(TableStyle([("BACKGROUND", (0, -1), (-1, -1), ACCENT_LIGHT)]))
    story.append(overview)
    for reason in incomplete:  # 자리 단위 "미완" 고지 (단계 문서 "하는 단계" 5)
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(f"미완 — {reason}", s["flag"]))
    story.append(Spacer(1, 7 * mm))

    story.append(_section_heading("2. 월별 비교·추이"))
    story.append(Spacer(1, 2 * mm))
    archived = load_archived_months()
    prev = _prev_month(base_month) if base_month else None
    has_comparison = bool(prev and prev in archived)

    def _comparison_flowable(width, font_size):
        """전월 대비 표 — 보관 데이터가 없으면 안내 문구를 대신 돌려준다."""
        if not has_comparison:
            return Paragraph(
                f"전월 대비 — 비교 대상 없음 (전월 {prev or '미상'} 보관 데이터 없음)", s["meta"])
        p = archived[prev]
        return _table(
            ["", f"당월 ({base_month})", f"전월 ({prev})", "증감"],
            [["총지출", f"{total_expense:,}원", f"{p.get('total_expense', 0):,}원",
              _fmt_delta(total_expense, p.get("total_expense", 0))],
             ["총수익", f"{total_income:,}원", f"{p.get('total_income', 0):,}원",
              _fmt_delta(total_income, p.get("total_income", 0))],
             ["순액", f"{net:,}원", f"{p.get('net', 0):,}원", _fmt_delta(net, p.get("net", 0))]],
            col_widths=_col_widths(0.14, 0.25, 0.25, 0.36, total=width),
            num_cols=(1, 2, 3), font_size=font_size,
        )

    # 같은 해 추이 — 보관된 월들 + 당월(집계 중인 데이터)
    year = base_month[:4] if base_month else None
    trend = sorted(
        [(m, d.get("total_expense", 0), d.get("total_income", 0))
         for m, d in archived.items() if year and m.startswith(year) and m != base_month]
        + ([(base_month, total_expense, total_income)] if base_month else [])
    )

    def _trend_table(width, font_size):
        return _table(
            ["월", "총지출", "총수익", "순액"],
            [[m, f"{exp:,}원", f"{inc:,}원", f"{inc - exp:,}원"] for m, exp, inc in trend],
            col_widths=_col_widths(0.22, 0.26, 0.26, 0.26, total=width),
            num_cols=(1, 2, 3), font_size=font_size,
        )

    if 2 <= len(trend) <= 4:
        # 막대그래프 구간 — 그래프를 왼쪽에 두고, 전월 대비·추이 표 두 개를
        # 그래프 오른쪽 빈 공간에 세로로 쌓아 섹션을 한 덩어리로 배치한다
        chart_width = CONTENT_WIDTH * 0.42
        gutter = 5 * mm
        table_width = CONTENT_WIDTH - chart_width - gutter
        right_column = [
            _comparison_flowable(table_width, 9),
            Spacer(1, 4 * mm),
            _trend_table(table_width, 9),
        ]
        left_column = [  # 그래프가 무엇의 추이인지 캡션으로 밝힌다
            Paragraph("월별 총지출", s["meta"]),
            _category_chart([(m, exp) for m, exp, _ in trend], width=chart_width),
        ]
        story.append(Table(
            [[left_column, right_column]],
            colWidths=[chart_width + gutter, table_width],
            style=TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]),
        ))
    else:
        story.append(_comparison_flowable(CONTENT_WIDTH, 10))
        if len(trend) >= 5:  # 월이 쌓이면 추세 읽기 좋은 꺾은선으로 전환
            story.append(Spacer(1, 4 * mm))
            story.append(_trend_table(CONTENT_WIDTH, 10))
            story.append(Spacer(1, 4 * mm))
            story.append(_trend_line_chart(trend))
        else:
            story.append(Spacer(1, 2 * mm))
            story.append(Paragraph(f"{year or '대상'}년 추이 — 표시할 다른 월 없음", s["meta"]))
    story.append(Spacer(1, 7 * mm))

    # ── 인사이트 계산 재료 — §3·§4·§5가 함께 쓴다 ──
    expense_rows = [r for r in ok_rows if r["지출"]]
    by_payer = dict(summary["by_payer"])
    foreign_rows = [r for r in ok_rows if (r.get("원거래통화") or "").strip()]  # 스키마 v1: 원거래통화 유무로 판별
    # 구분별 전월 대비의 데이터 원천은 보관 집계의 by_payer (interface-spec.md 산출물 양식)
    prev_by_payer = dict(archived[prev].get("by_payer", []) or []) if has_comparison else {}

    story.append(_section_heading("3. 종합 분석·인사이트"))
    story.append(Spacer(1, 2 * mm))
    common_lines = []
    line = f"총지출 {total_expense:,}원 · 총수익 {total_income:,}원 · 순액 {net:,}원"
    if has_comparison:
        line += f" — 지출 전월 대비 {_fmt_delta(total_expense, archived[prev].get('total_expense', 0))}"
    common_lines.append(line)
    if summary["by_category"]:
        top_cat, top_amt = summary["by_category"][0]
        common_lines.append(
            f"최대 지출 카테고리는 {top_cat} — 전체 지출의 {(top_amt / total_expense * 100 if total_expense else 0):.1f}%")
    if summary["top_spenders"]:
        t = summary["top_spenders"][0]
        common_lines.append(f"최대 단일 지출은 {t['날짜']} {t['결제처']} {t['지출']:,}원")
    if foreign_rows:
        foreign_sum = sum((r["지출"] or r["수익"] or 0) for r in foreign_rows)
        common_lines.append(f"해외결제 {len(foreign_rows)}건 · 원화 환산 {foreign_sum:,}원 — 명세는 §6 참조")
    misc_expense = sum(v for p, v in by_payer.items() if p not in ("개인결제", "법인결제"))
    if misc_expense:
        common_lines.append(f"결제구분 미상 지출 {misc_expense:,}원 — 구분 확인 필요")
    if flagged_rows:
        n_reject = sum(1 for r in flagged_rows if r.get("type") != "확인 요청")
        n_review = len(flagged_rows) - n_reject
        common_lines.append(
            f"확인 필요 {len(flagged_rows)}건 (반려 {n_reject} · 확인 요청 {n_review}) — 목록은 §7 참조")
    else:
        common_lines.append("확인 필요 항목 없음")
    story.append(_insight_box(common_lines, s["insight"]))

    band_items = [(p, v) for p, v in by_payer.items() if v > 0]
    if len(band_items) >= 2:  # 구성비는 조각이 둘 이상일 때만 의미 있다
        story.append(Spacer(1, 4 * mm))
        story.append(KeepTogether([  # 소제목이 페이지 끝에 홀로 남지 않게 그래프와 묶는다
            Paragraph("개인/법인 구성비", s["body"]),
            Spacer(1, 1 * mm),
            _share_band(band_items),
        ]))
    methods = sorted(summary["by_method"], key=lambda kv: kv[1], reverse=True)
    if methods:
        story.append(Spacer(1, 4 * mm))
        # 막대 라벨이 금액·비중을 다 보여주므로 같은 수치의 표는 싣지 않는다 (중복 게재 금지)
        story.append(KeepTogether([
            Paragraph("결제수단별 합계", s["body"]),
            _hbar_chart(methods, total=total_expense),
        ]))
    if summary["top_spenders"]:
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("주요 지출 Top 10 (지출 행만)", s["body"]))
        story.append(Spacer(1, 1 * mm))
        story.append(_table(
            ["날짜", "결제처", "금액"],
            [[r["날짜"], Paragraph(r["결제처"], s["cell"]), f"{r['지출']:,}원"] for r in summary["top_spenders"]],
            col_widths=_col_widths(0.23, 0.54, 0.23), num_cols=(2,),
        ))
    story.append(Spacer(1, 7 * mm))

    def _payer_section(payer, corporate=False):
        """구분(개인/법인) 하나의 분석·인사이트 묶음 — 카테고리 파이 + 규칙 기반 요약문.
        카테고리 세부 표는 파이 범례(금액·비중)가 대신한다 (중복 게재 금지)."""
        rows_p = [r for r in expense_rows if (r["결제자"] or "").strip() == payer]
        subtotal = by_payer.get(payer, 0)
        if not rows_p:
            return [Paragraph(f"해당 없음 ({payer} 건 없음)", s["body"])]
        detail = summary["by_payer_category"].get(payer, [])
        out = [_pie_chart(_group_top(detail)), Spacer(1, 3 * mm)]

        lines = []
        line = f"{payer} 지출 {subtotal:,}원 — 전체 지출의 {(subtotal / total_expense * 100 if total_expense else 0):.1f}%"
        if payer in prev_by_payer:
            line += f" · 전월 대비 {_fmt_delta(subtotal, prev_by_payer[payer])}"
        lines.append(line)
        if detail:
            top_cat, top_amt = detail[0]
            lines.append(f"최대 카테고리는 {top_cat} — {payer} 지출의 {(top_amt / subtotal * 100 if subtotal else 0):.1f}%")
        biggest = max(rows_p, key=lambda r: r["지출"])
        lines.append(f"최고 단일 건은 {biggest['날짜']} {biggest['결제처']} {biggest['지출']:,}원")
        merchants = {}
        for r in rows_p:
            name = (r["결제처"] or "").strip()
            if name:
                entry = merchants.setdefault(name, [0, 0])
                entry[0] += 1
                entry[1] += r["지출"]
        repeats = sorted(((name, cnt, amt) for name, (cnt, amt) in merchants.items() if cnt >= 2),
                         key=lambda item: item[2], reverse=True)[:2]
        if repeats:
            lines.append("같은 결제처 반복 결제: " + " · ".join(f"{name} {cnt}회 {amt:,}원" for name, cnt, amt in repeats))
        if corporate:
            # 법인은 지출 통제 관점 — 주말 결제는 0건이어도 확인 결과로 명시한다
            weekend = [r for r in rows_p if _is_weekend(r.get("날짜"))]
            if weekend:
                lines.append(f"주말 결제 {len(weekend)}건 · {sum(r['지출'] for r in weekend):,}원 — 업무 관련성 확인 권장")
            else:
                lines.append("주말 결제 없음")
            if any("verify3_result" in r for r in ok_rows):  # 부정 사용 감지 토글을 켠 실행에만 결과가 있다
                n_review = sum(1 for r in flagged_rows if r.get("type") == "확인 요청")
                lines.append(f"부정 사용 감지 확인 요청 {n_review}건" + (" — 목록은 §7 참조" if n_review else ""))
        out.append(_insight_box(lines, s["insight"]))
        return out

    story.append(_section_heading("4. 법인결제 분석·인사이트"))
    story.append(Spacer(1, 2 * mm))
    story.extend(_payer_section("법인결제", corporate=True))
    story.append(Spacer(1, 7 * mm))

    story.append(_section_heading("5. 개인결제 분석·인사이트"))
    story.append(Spacer(1, 2 * mm))
    story.extend(_payer_section("개인결제"))
    story.append(Spacer(1, 7 * mm))

    story.append(_section_heading("6. 해외결제 명세"))
    story.append(Spacer(1, 2 * mm))
    if foreign_rows:
        story.append(_table(
            ["날짜", "결제처", "원거래", "원화 환산"],
            [[r["날짜"], Paragraph(r["결제처"], s["cell"]), f"{r['원거래금액']} {r['원거래통화']}",
              f"{(r['지출'] or r['수익'] or 0):,}원"]
             for r in foreign_rows],
            col_widths=_col_widths(0.18, 0.34, 0.24, 0.24), num_cols=(3,),
        ))
    else:
        story.append(Paragraph("해당 없음 (해외결제 건 없음)", s["body"]))
    story.append(Spacer(1, 7 * mm))

    story.append(_section_heading("7. 확인 필요 항목"))
    story.append(Spacer(1, 2 * mm))
    if flagged_rows:
        table = _table(
            ["transaction_id", "날짜", "결제처", "유형", "사유"],
            # 긴 결제처명이 이웃 열을 침범하지 않게 결제처도 Paragraph로 줄바꿈시킨다
            [[r.get("transaction_id", ""), r["날짜"], Paragraph(r["결제처"], s["cell"]), r.get("type", "반려"),
              Paragraph(r["reason"], s["flag"])] for r in flagged_rows],
            col_widths=_col_widths(0.18, 0.14, 0.18, 0.10, 0.40),
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), FLAG_COLOR),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PAGE_BG, FLAG_LIGHT]),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("확인 필요 항목 없음", s["body"]))

    doc.build(story, onFirstPage=_paint_page_background, onLaterPages=_paint_page_background)


def run(month=None, fraud_check=False):
    """xlsx·pdf를 실제로 만들고, 화면 미리보기·지휘 보고에 쓸 결과를 함께 돌려준다.

    month: 지휘가 전달하는 대상 월(YYYY-MM, 실행 파라미터) — 리포트 개요에 싣는다.
    미전달(단독 실행·화면 경로)이면 데이터에서 파생해 추정 표기한다.
    fraud_check: 이번 실행의 부정 사용 감지 토글 값 — load_transactions()로 그대로 전달한다.
    """
    if not os.path.exists(REFINE_CSV):
        failed_envelope = build_envelope(
            0, [], [], failed=True,
            message=f"가공 산출물 없음: {os.path.relpath(REFINE_CSV, REPO_ROOT)} — 취합 불가",
        )
        return {"ok_rows": [], "flagged_rows": [], "summary": summarize([]), "envelope": failed_envelope}

    transactions, incomplete = load_transactions(fraud_check=fraud_check)
    total = len(transactions)
    if total == 0:
        empty_envelope = build_envelope(0, [], [], message="확인 대상 없음")
        return {"ok_rows": [], "flagged_rows": [], "summary": summarize([]), "envelope": empty_envelope}

    ok_rows, flagged_rows = split_transactions(transactions)
    summary = summarize(ok_rows)
    try:
        build_ledger(ok_rows, XLSX_PATH)
        build_report(ok_rows, flagged_rows, summary, PDF_PATH, month=month, incomplete=incomplete)
    except Exception as e:
        envelope = build_envelope(total, ok_rows, flagged_rows, failed=True, message=f"산출물 생성 실패: {e}")
        return {"ok_rows": ok_rows, "flagged_rows": flagged_rows, "summary": summary, "envelope": envelope}

    envelope = build_envelope(total, ok_rows, flagged_rows,
                              message=" / ".join(incomplete), incomplete=incomplete)
    return {"ok_rows": ok_rows, "flagged_rows": flagged_rows, "summary": summary, "envelope": envelope}


def main():
    result = run()
    envelope = result["envelope"]
    if envelope["status"] == "empty":
        print("확인 대상 없음")
        return

    print(f"엑셀 회계장부: {XLSX_PATH}")
    print(f"PDF 결산 리포트: {PDF_PATH}")
    print(f"장부 반영 {len(result['ok_rows'])}건 · 확인 필요 {len(result['flagged_rows'])}건")
    print("지휘에게 보내는 단계 결과 보고:")
    print(json.dumps(envelope, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
